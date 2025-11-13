from __future__ import annotations
from collections import defaultdict
from typing import Dict, Tuple, Any, Optional, List
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from .config import (
    MASTER_COLUMNS,
    MASTER_MONTH_FORMAT,
    MONTH_NAME_MAP,
    BARGE_COLUMNS,
    RunOptions,
    fmt_money,
    COMPOUND_FLF_MODE,
    FORCE_MONTH_FROM_HEADER,
    MASTER_SHEET_NAMES,
)
from .mapping import canonicalize, MASTER_KEYS


# ------------------------- helpers -------------------------
def excel_col_to_idx(col_letter: str) -> int:
    return column_index_from_string(col_letter)


def month_string_to_master_key(month_cell: Any, year: int) -> str:
    """
    Terima 'Aug', 'AUG', '1 Aug', 'Aug 2025', or datetime -> 'Aug-25'
    """
    if pd.isna(month_cell):
        raise ValueError("Month cell is empty")

    if isinstance(month_cell, (datetime, pd.Timestamp)):
        dt = pd.Timestamp(month_cell).to_pydatetime().replace(year=year)
        return dt.strftime(MASTER_MONTH_FORMAT)

    s = str(month_cell).strip()
    letters = "".join(re.findall(r"[A-Za-z]+", s))[:3].lower()
    m = MONTH_NAME_MAP.get(letters)
    if not m:
        raise ValueError(f"Cannot parse month value: {s!r}")
    dt = datetime(year, m, 1)
    return dt.strftime(MASTER_MONTH_FORMAT)


BARGE_HEADER_ALIASES = {
    "No": ["NO", "NO."],
    "Month": ["MONTH", "MON"],
    "Status": ["STATUS"],
    "ActualLoaded": ["QTY ACTUAL LOADED", "ACTUAL LOADED", "QTY ACTUAL LOADING", "ACTUAL LOADING"],
    "LoadingFacilities": ["LOADING FACILITIES", "LOADING FACILITY"],
    "FLFNominate": ["FLF NOMINATE", "FLF/FC NOMINATE", "FLF FC NOMINATE"],
}

REQUIRED_BARGE_KEYS = {"No", "Month", "Status", "ActualLoaded", "LoadingFacilities", "FLFNominate"}


def _detect_barge_columns(df, hint_row: Optional[int] = None) -> Dict[str, int]:
    """
    Cari baris header terbaik di jendela sekitar hint_row.
    Fallback: awal sheet (0..400), lalu lebih dalam (0..1200), lalu mapping lama dari config.
    Mengembalikan map {logical_name: col_index_0_based} dan kunci tambahan '_header_row'.
    """
    # def _normalize_block(block: pd.DataFrame) -> pd.DataFrame:
    #     return block.fillna("").astype(str).applymap(lambda s: s.strip().upper())

    def _normalize_block(block: pd.DataFrame) -> pd.DataFrame:
        return (
        block
        .fillna("")
        .astype(str)
        .applymap(lambda s: s.strip().upper())
    )

    def best_map_in(range_df: pd.DataFrame) -> Tuple[Dict[str, int], int]:
        # return (map_kolom, indeks_baris_dalam_range)
        norm = _normalize_block(range_df)
        best_row = -1
        best_map: Dict[str, int] = {}
        for r, row in norm.iterrows():
            m: Dict[str, int] = {}
            for c, txt in enumerate(row.tolist()):
                if not txt:
                    continue
                for key, aliases in BARGE_HEADER_ALIASES.items():
                    if key in m:
                        continue
                    # utamakan exact match; jika tidak ada, izinkan contains
                    if any(txt == a or a in txt for a in aliases):
                        m[key] = c
            if len(m) > len(best_map):
                best_map, best_row = m, r
                if REQUIRED_BARGE_KEYS.issubset(best_map.keys()):
                    break
        return best_map, best_row

    # Susun jendela scan: sekitar Start Row → awal sheet → lebih dalam
    windows: List[Tuple[int, int]] = []
    n = len(df)
    if hint_row is not None:
        windows.append((max(0, hint_row - 50), min(n, hint_row + 200)))
    windows += [(0, min(400, n)), (0, min(1200, n))]

    for r0, r1 in windows:
        sub = df.iloc[r0:r1]
        m, row_in_sub = best_map_in(sub)
        if not m:
            continue
        miss = REQUIRED_BARGE_KEYS - set(m)
        if not miss:
            m["_header_row"] = r0 + row_in_sub  # info tambahan
            return m

    # Fallback terakhir: mapping huruf kolom lama
    from .config import BARGE_COLUMNS
    from openpyxl.utils import column_index_from_string as __col
    m2: Dict[str, int] = {}
    for name, letter in BARGE_COLUMNS.items():
        idx = __col(letter) - 1
        if idx < df.shape[1]:
            m2[name] = idx
    miss = REQUIRED_BARGE_KEYS - set(m2)
    if miss:
        raise ValueError(f"Tidak bisa deteksi kolom Barge (kurang: {', '.join(sorted(miss))}).")
    m2["_header_row"] = 0
    return m2


def read_barge_rows(opts: RunOptions) -> tuple[pd.DataFrame, str | None]:
    df = pd.read_excel(opts.barge_path, sheet_name=opts.barge_sheet, header=None, engine="openpyxl")

    # >>> deteksi kolom berdasarkan header (pakai hint dari Start Row)
    keep_idx = _detect_barge_columns(df, hint_row=opts.start_row - 1)
    header_row = keep_idx.pop("_header_row", -1)  # buang kunci helper, simpan nilainya

    view = df.iloc[:, list(keep_idx.values())].copy()
    view.columns = list(keep_idx.keys())

    # start/end slice
    start0 = max(opts.start_row - 1, 0)

    # kalau user start di atas header, geser ke bawah header
    if header_row >= 0 and start0 <= header_row:
        start0 = header_row + 1

    end0 = start0 + opts.row_count if opts.row_count > 0 else None

    # deteksi header month (scan beberapa baris ke depan)
    header_month_key = None
    try:
        for r in range(start0, min((start0 + 50), len(df))):
            for key in ("Month", "No"):
                cidx = keep_idx.get(key)
                if cidx is None:
                    continue
                cand = df.iat[r, cidx]
                if cand is None or (isinstance(cand, float) and pd.isna(cand)):
                    continue
                try:
                    header_month_key = month_string_to_master_key(cand, opts.target_year)
                    break
                except Exception:
                    continue
            if header_month_key:
                break
    except Exception:
        header_month_key = None

    # slice baris yang diproses
    view = view.iloc[start0:end0]

    # ==== FILTER STATUS dengan "pinjam status atas" untuk baris data yang kosong statusnya ====
    if opts.only_completed:
        needed = ("Status", "FLFNominate", "LoadingFacilities", "ActualLoaded")
        if all(col in view.columns for col in needed):
            # normalisasi status
            st_raw = (
                view["Status"].astype(str).str.strip().str.upper()
                .replace({"NAN": ""})
            )

            # baris terlihat seperti data (bukan subtotal) jika punya tujuan/fasilitas
            looks_data = (
                view["FLFNominate"].astype(str).str.strip().ne("") |
                view["LoadingFacilities"].astype(str).str.strip().ne("")
            )

            # helper: cek apakah kolom ActualLoaded berisi angka
            def _is_number(x):
                try:
                    s = str(x).strip().replace("\u00A0", " ").replace(" ", "")
                    if "." in s and "," in s:
                        s = s.replace(".", "").replace(",", ".")
                    elif "," in s:
                        part = s.split(",")[-1]
                        s = s.replace(",", "") if len(part) == 3 else s.replace(",", ".")
                    float(s)
                    return True
                except Exception:
                    return False

            has_amount = view["ActualLoaded"].apply(_is_number)

            # hanya pinjam status dari baris atas kalau: status kosong + ini baris data + ada nilai
            carry_ok = st_raw.eq("") & looks_data & has_amount
            st_eff = st_raw.mask(carry_ok, st_raw.shift(1))

            # pakai status efektif untuk filter
            view = view[st_eff.isin({"COMPLETE", "COMPLETED"})]
        else:
            # fallback kalau ada kolom yang kurang: pakai filter lama
            status = view.get("Status", "").astype(str).str.strip().str.upper()
            view = view[status.isin({"COMPLETE", "COMPLETED"})]


    return view, header_month_key


def aggregate_barge(
    view: pd.DataFrame,
    target_year: int,
    header_month_key: str | None = None
) -> Dict[str, Dict[str, float]]:
    """
    Kembalikan dict: { 'Aug-25': {'Apollo': total, 'Zeus': total, ... }, ... }
    Aturan gabungan:
      - COMPOUND_FLF_MODE == "split": bagi rata ke target valid
      - "first": hanya target pertama
      - selain itu: duplicate (nilai penuh ke semua target valid)
    """

    # --- optional manual override dari config (kalau kamu tambahkan)
    try:
        from .config import FORCE_MONTH_OVERRIDE  # type: ignore
        manual_override = (FORCE_MONTH_OVERRIDE or "").strip()
    except Exception:
        manual_override = ""

    # prioritas bulan: manual override > header dari baris start (jika FORCE_MONTH_FROM_HEADER=True) > per-baris
    forced_key = manual_override or (header_month_key if FORCE_MONTH_FROM_HEADER else "")

    by_month: Dict[str, Dict[str, float]] = defaultdict(lambda: defaultdict(float))

    for _, row in view.iterrows():
        # tentukan month key
        try:
            if forced_key:
                month_key = forced_key
            else:
                month_key = month_string_to_master_key(row.get("Month"), target_year)
        except Exception:
            continue

        raw_flf = str(row.get("FLFNominate", "")).strip()
        actual = row.get("ActualLoaded", 0)

        # parse angka
        try:
            val = _to_float(actual)
        except Exception:
            continue

        # mapping FLF
        targets = [t for t in canonicalize(raw_flf) if t in MASTER_KEYS]
        if not targets:
            continue

        if COMPOUND_FLF_MODE == "split" and len(targets) > 1:
            share = val / len(targets)
            for t in targets:
                by_month[month_key][t] += share
        elif COMPOUND_FLF_MODE == "first":
            by_month[month_key][targets[0]] += val
        else:
            for t in targets:
                by_month[month_key][t] += val

    return by_month


def _to_float(x) -> float:
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace("\u00A0", " ")  # non-breaking space → space
    s = s.replace(" ", "")                     # buang spasi
    # Jika ada titik & koma: anggap titik = pemisah ribuan, koma = desimal
    if "." in s and "," in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        # Hanya koma → tentukan apakah ribuan atau desimal
        if "," in s:
            part = s.split(",")[-1]
            # jika 3 digit di belakang koma, ini ribuan (1,234) ⇒ hilangkan koma
            if len(part) == 3:
                s = s.replace(",", "")
            else:
                s = s.replace(",", ".")
        # Hanya titik → biarkan (asumsikan US style 1234.56)
    s = re.sub(r"[^0-9.]", "", s)  # sisakan digit & titik
    return float(s) if s else 0.0


def find_master_row(ws, month_key: str) -> int | None:
    month_col = excel_col_to_idx(MASTER_COLUMNS["Month/Year"])
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(r, month_col).value
        if cell is None:
            continue
        if isinstance(cell, str):
            if cell.strip().lower() == month_key.lower():
                return r
        elif isinstance(cell, datetime):
            if cell.strftime(MASTER_MONTH_FORMAT).lower() == month_key.lower():
                return r
    return None

def apply_to_master(
    master_path,
    sheet_name: str,
    totals: Dict[str, Dict[str, float]],
    dry_run: bool = False,
    clear_before_write: bool = True,
    clear_value: float | None = 0.0,  # ← biarin; nanti kita panggil dengan None
) -> list:

    """
    Tulis hasil agregasi ke workbook master.
    Jika clear_before_write=True: kosongkan semua kolom FLF di baris bulan tsb dulu,
    lalu tulis angka baru (REPLACE).
    """
    wb = load_workbook(master_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} not found in master.")

    ws = wb[sheet_name]
    logs: list = []

    # indeks semua kolom FLF kecuali kolom Month/Year
    col_index = {k: excel_col_to_idx(v) for k, v in MASTER_COLUMNS.items() if k != "Month/Year"}

    def _as_float_or_zero(x) -> float:
        try:
            return float(str(x).replace(",", ""))
        except Exception:
            return 0.0

    for month_key, flf_totals in sorted(totals.items()):
        r = find_master_row(ws, month_key)
        if r is None:
            logs.append(f"[WARN] Month '{month_key}' not found in master sheet '{sheet_name}'. Skipped.")
            continue

        # --- CLEAR row bulan ini dulu (opsional)
        if clear_before_write:
            for flf_name, c in col_index.items():
                cell = ws.cell(r, c)
                old_val = _as_float_or_zero(cell.value)
                new_v = None if clear_value is None else float(clear_value)  # ← ini yang bikin kosong
                
                if not dry_run:
                    cell.value = new_v
                logs.append(f"[{month_key}] {flf_name}: cleared (was {fmt_money(old_val)})")

        # lalu tulis angka baru
        for flf_name, new_val in flf_totals.items():
            c = col_index.get(flf_name)
            if not c:
                logs.append(f"[WARN] Unknown FLF column '{flf_name}' in master. Skipped.")
                continue
            cell = ws.cell(r, c)
            old_val = _as_float_or_zero(cell.value)
            logs.append(f"[{month_key}] {flf_name}: set to {fmt_money(new_val)} (was {fmt_money(old_val)})")
            if not dry_run:
                cell.value = float(new_val)

    if not dry_run:
        wb.save(master_path)
        logs.append(f"[OK] Saved changes to: {master_path}")
    else:
        logs.append("[DRY-RUN] No changes written.")
    return logs


def run_pipeline(opts: RunOptions, progress=lambda *_: None) -> Tuple[Dict[str, Dict[str, float]], list]:
    progress(f"Reading barge '{opts.barge_sheet}' rows {opts.start_row}..{opts.start_row + max(opts.row_count-1,0)}")
    view, header_month_key = read_barge_rows(opts)
    progress(f"Rows after filter: {len(view)} | header_month={header_month_key or '-'}")

    totals = aggregate_barge(view, opts.target_year, header_month_key=header_month_key)
    n_months = len(totals); n_cells = sum(len(x) for x in totals.values())
    progress(f"Aggregated months: {n_months} | cells to write: {n_cells}")

    if n_cells == 0:
        warn = "[WARN] No totals aggregated — cek Status filter, parsing 'ActualLoaded', mapping 'FLFNominate', atau format bulan."
        progress(warn)
        return totals, [warn]
    
    sheet_name = MASTER_SHEET_NAMES.get(opts.target_year, str(opts.target_year))
    progress(f"Writing to master sheet '{sheet_name}' (dry_run={opts.dry_run}) …")
    logs = apply_to_master(
        opts.master_path,
        sheet_name,
        totals,
        dry_run=opts.dry_run,
        clear_before_write=opts.clear_before_write,
        clear_value=None,  # ← KOSONGKAN sel saat clear
    )

    progress("Done.")
    return totals, logs
