# app/interface.py  -- PyQt5 (preflight + baca opsi dari UI utama, SELALU APPLY)
import os
import ui.resources_rc  # penting: register Qt resources (:/img/2.jpg, :/img/power.ico)
from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from app.config import (
    RunOptions, DEFAULT_BARGE_SHEET, MASTER_SHEET_NAMES, BARGE_COLUMNS,
)
from app.main_logic import run_pipeline
from app.mapping import MASTER_KEYS

# ingat folder terakhir
_settings = QtCore.QSettings("YourCompany", "YourApp")

# =========================================================
# Utils
# =========================================================
def _browse_into(parent: QtWidgets.QWidget,
                 line_edit: QtWidgets.QLineEdit,
                 title: str,
                 filters: str = "Excel (*.xlsx);;All Files (*.*)") -> str:
    start_dir = _settings.value(
        "lastDir",
        QtCore.QStandardPaths.writableLocation(QtCore.QStandardPaths.DocumentsLocation)
        or os.path.expanduser("~")
    )
    path, _ = QtWidgets.QFileDialog.getOpenFileName(parent, title, start_dir, filters)
    if path:
        line_edit.setText(path)
        _settings.setValue("lastDir", os.path.dirname(path))
    return path


def _populate_sheet_combo(ui, barge_path: str):
    """Isi combobox sheet bila ada di UI utama."""
    if not hasattr(ui, "comboSheet") or not barge_path:
        return
    try:
        wb = load_workbook(barge_path, read_only=True)
        ui.comboSheet.clear()
        ui.comboSheet.addItems(wb.sheetnames)
        if DEFAULT_BARGE_SHEET in wb.sheetnames:
            ui.comboSheet.setCurrentText(DEFAULT_BARGE_SHEET)
    except Exception:
        pass  # biarkan isi sebelumnya


def _preflight_barge(barge_path: str, sheet: str, start_row: int):
    """Validasi dasar agar tidak out-of-bounds."""
    if not os.path.exists(barge_path):
        raise ValueError("Barge workbook tidak ditemukan.")

    wb = load_workbook(barge_path, read_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Worksheet '{sheet}' tidak ada. Pilih sheet yang benar.")

    ws = wb[sheet]
    max_row, max_col = ws.max_row, ws.max_column

    # pastikan kolom yang dibutuhkan ada
    req_cols = {k: column_index_from_string(v) for k, v in BARGE_COLUMNS.items()}
    missing = [k for k, idx in req_cols.items() if idx > max_col]
    if missing:
        raise ValueError(
            "Template barge tidak sesuai. Kolom berikut tidak ditemukan: "
            + ", ".join(missing) + f" (max column={max_col})."
        )

    if start_row < 1 or start_row > max_row:
        raise ValueError(
            f"Start Row {start_row} di luar jangkauan (jumlah baris: {max_row})."
        )


def _options_summary_text(opts: RunOptions) -> str:
    """Ringkasan opsi dalam plain text untuk ditaruh di log StartDialog."""
    lines = [
        f"Master        : {opts.master_path}",
        f"Barge         : {opts.barge_path}",
        f"Year          : {opts.target_year}",
        f"Barge Sheet   : {opts.barge_sheet}",
        f"Start Row     : {opts.start_row}",
        f"Row Count     : {opts.row_count}",
        f"Only Completed: {'Yes' if opts.only_completed else 'No'}",
        f"Dry Run       : {'Yes' if opts.dry_run else 'No'}",
        "",
        "Click the power button to start…",
    ]
    return "\n".join(lines)


def _show_error_dialog(parent: QtWidgets.QWidget, message: str):
    """Dialog error modern dengan 1 tombol OK."""
    mb = QtWidgets.QMessageBox(parent)
    mb.setIcon(QtWidgets.QMessageBox.Critical)
    mb.setWindowTitle("Error")
    mb.setText("An error occurred during processing:")
    mb.setInformativeText(message)
    mb.setStandardButtons(QtWidgets.QMessageBox.Ok)
    mb.setDefaultButton(QtWidgets.QMessageBox.Ok)
    mb.setWindowFlags(mb.windowFlags() & ~Qt.WindowContextHelpButtonHint)
    mb.setStyleSheet("""
        QMessageBox { background:#eef2f7; }
        QLabel { color:#111827; font:600 10pt "Segoe UI"; }
        QLabel#qt_msgbox_label { font:700 11pt "Segoe UI"; }
        QPushButton {
          background:#0d6efd; color:#fff; border:none; border-radius:8px;
          padding:6px 16px; min-width:70px; font:600 10pt "Segoe UI";
        }
        QPushButton:hover  { background:#0b5ed7; }
        QPushButton:pressed{ background:#0a58ca; }
    """)
    mb.exec_()


def _show_result_dialog(parent: QtWidgets.QWidget, *, is_dry_run: bool):
    """Dialog sukses minimalis (selaras mainWindow)."""
    dlg = QtWidgets.QDialog(parent)
    dlg.setWindowTitle("Hasil Proses")
    dlg.resize(520, 150)
    dlg.setWindowFlags(dlg.windowFlags() & ~Qt.WindowContextHelpButtonHint)

    root = QtWidgets.QVBoxLayout(dlg)
    root.setContentsMargins(18, 16, 18, 16)
    root.setSpacing(14)

    card = QtWidgets.QFrame(objectName="card")
    lay = QtWidgets.QHBoxLayout(card)
    lay.setContentsMargins(14, 12, 14, 12)
    lay.setSpacing(12)

    icon = QtWidgets.QLabel("✅")
    icon.setFont(QtGui.QFont("Segoe UI Emoji", 22))
    msg = QtWidgets.QLabel(
        "Perubahan telah diterapkan pada file master."
        if not is_dry_run else
        "Simulasi selesai (DRY-RUN).\nTidak ada perubahan yang disimpan."
    )
    msg.setObjectName("msg")
    msg.setWordWrap(True)

    lay.addWidget(icon, 0, Qt.AlignTop)
    lay.addWidget(msg, 1)
    root.addWidget(card)

    btns = QtWidgets.QHBoxLayout(); btns.addStretch(1)
    ok = QtWidgets.QPushButton("OK", objectName="btnPrimary")
    ok.setMinimumHeight(34); ok.clicked.connect(dlg.accept)
    btns.addWidget(ok)
    root.addLayout(btns)

    dlg.setStyleSheet("""
      QDialog { background:#f3f4f6; }
      QFrame#card { background:#ffffff; border:1px solid #e5e7eb; border-radius:12px; }
      QLabel#msg { color:#0f172a; font:600 10pt "Segoe UI"; }
      QPushButton#btnPrimary {
        background:#0d6efd; color:#fff; border:none; border-radius:10px;
        padding:6px 22px; font:700 10pt "Segoe UI";
      }
      QPushButton#btnPrimary:hover  { background:#0b5ed7; }
      QPushButton#btnPrimary:pressed{ background:#0a58ca; }
    """)
    dlg.exec_()

# =========================================================
# Background worker + StartDialog
# =========================================================
class _Worker(QtCore.QThread):
    """Thread ringan untuk menjalankan callable `runner` (return (totals, logs))."""
    progress = QtCore.pyqtSignal(str)
    finished = QtCore.pyqtSignal(object, object)   # totals, logs
    failed = QtCore.pyqtSignal(str)

    def __init__(self, runner, parent=None):
        super().__init__(parent)
        self._runner = runner

    def run(self):
        try:
            self.progress.emit("Starting...")
            totals, logs = self._runner()
            self.finished.emit(totals, logs)
        except Exception as e:
            self.failed.emit(str(e))


class StartDialog(QtWidgets.QDialog):
    """
    Dialog 'Start Automation Process' dengan tombol power.
    Klik power => jalankan _Worker, tampilkan log. Selesai -> dialog sukses.
    """
    def __init__(self, parent, runner, *, is_dry_run=False, preface_text: str = ""):
        super().__init__(parent)
        self.setObjectName("StartDialog")
        self.setWindowTitle("Start Automation Process")
        self.setModal(True)
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        self.resize(720, 480)

        self._runner = runner
        self._is_dry_run = is_dry_run

        # === Background & styles ===
        self.setStyleSheet("""
            QDialog#StartDialog {
                border-image: url(:/img/2.jpg) 0 0 0 0 stretch stretch;
            }
            QLabel#title {
                font: 600 28pt "Segoe UI";
                color: rgb(255,255,255);
            }
            QTextEdit {
                background: rgba(255,255,255,0.92);
                border: 1px solid #e5e7eb;
                border-radius: 12px;
                color: #0f172a;
                font: 10pt "Segoe UI";
            }
            QPushButton.primary {
                background:#0d6efd; color:#fff; border:none; border-radius:10px;
                padding:6px 22px; font:700 10pt "Segoe UI";
            }
            QPushButton.primary:hover  { background:#0b5ed7; }
            QPushButton.primary:pressed{ background:#0a58ca; }
            QPushButton.primary:disabled{ background:#9ca3af; }
        """)

        # ================== LAYOUT ROOT ==================
        self._root = QtWidgets.QVBoxLayout(self)
        self._root.setContentsMargins(24, 20, 24, 20)
        self._root.setSpacing(16)

        # Header
        title = QtWidgets.QLabel("Start Automation Process")
        title.setObjectName("title")                 # gunakan #title (bukan class)
        title.setAlignment(Qt.AlignHCenter)
        self._root.addWidget(title)

        # Tombol Power — polos (pakai ikon yang sudah ada backgroundnya)
        power_btn = QtWidgets.QPushButton()
        self._power_btn = power_btn
        power_btn.setFlat(True)                               # hilangin frame/button chrome
        power_btn.setCursor(Qt.PointingHandCursor)
        power_btn.setAutoDefault(False); power_btn.setDefault(False)
        power_btn.setIcon(QtGui.QIcon(":/img/power.ico"))     # ikon hasil edit kamu
        power_btn.setIconSize(QtCore.QSize(128, 128))         # sesuaikan dengan ukuran ikon
        power_btn.setFixedSize(128, 128)                      # biar area klik pas
        power_btn.setEnabled(True)
        power_btn.setStyleSheet("QPushButton{background:transparent;border:none;}")


        hb = QtWidgets.QHBoxLayout()
        hb.addStretch(1); hb.addWidget(power_btn); hb.addStretch(1)
        self._root.addLayout(hb)

        # Log area (preface plain text)
        self.log = QtWidgets.QTextEdit()
        self.log.setObjectName("textLog")
        self.log.setReadOnly(True)
        self.log.setAcceptRichText(True)
        if preface_text:
            self.log.setPlainText(preface_text)
        else:
            self.log.setPlaceholderText("Click the power button to start…")
        self._root.addWidget(self.log, 1)

        # Footer
        f = QtWidgets.QHBoxLayout(); f.addStretch(1)
        self.btnEnd = QtWidgets.QPushButton("End")
        self.btnEnd.setEnabled(False)
        self.btnEnd.setCursor(Qt.PointingHandCursor)
        self.btnEnd.setMinimumHeight(36)
        self.btnEnd.setProperty("class", "primary")
        f.addWidget(self.btnEnd)
        self._root.addLayout(f)

        # Background worker & signals
        self.worker = _Worker(self._runner, self)
        self.worker.progress.connect(self._append)
        self.worker.finished.connect(self._on_done)
        self.worker.failed.connect(self._on_failed)
        power_btn.clicked.connect(self._on_start)
        self.btnEnd.clicked.connect(self.accept)

    # --- slots/helpers
    def _append(self, text: str):
        self.log.append(text)

    def _append_html(self, html: str):
        cur = self.log.textCursor()
        cur.movePosition(cur.End)
        cur.insertHtml(html + "<br/>")
        self.log.setTextCursor(cur)
        self.log.ensureCursorVisible()

    def _ok(self, text: str):
        self._append_html(f'<span style="color:#166534;">{text}</span>')

    def _err(self, text: str):
        self._append_html(f'<span style="color:#b91c1c;"><b>{text}</b></span>')

    def _on_start(self):
        self.log.clear()
        self._append("Processing data...")
        # disable semua tombol kecuali 'End'
        for w in self.findChildren(QtWidgets.QPushButton):
            if w is not self.btnEnd:
                w.setEnabled(False)
        self.worker.start()

    def _on_done(self, totals, logs):
        # Ringkasan (pakai HTML untuk highlight)
        self._append_html("<br/><b>Summary</b>")
        try:
            self._append_html(self._format_summary(totals or {}))
        except Exception as e:
            self._err(f"Gagal membuat ringkasan: {e}")

        # Detail logs
        if logs:
            self._append_html("<br/><b>Details</b>")
            for line in logs:
                s = str(line)
                up = s.upper()
                if "ERROR" in up or "FAILED" in up:
                    self._err(s)
                elif s.strip().startswith("[OK]") or "SAVED CHANGES" in up or "SUCCESS" in up:
                    self._ok(s)
                else:
                    self._append_html(s)

        self._append_html("<br/><b>Execution completed.</b>")
        self._ok("Automation completed successfully!")
        self.btnEnd.setEnabled(True)

        # Pop-up sukses
        _show_result_dialog(self, is_dry_run=self._is_dry_run)

    def _on_failed(self, message: str):
        _show_error_dialog(self, message)
        self.btnEnd.setEnabled(True)

    # ---- Ringkasan untuk tampil di log -----------------------------------------
    def _format_summary(self, totals: dict) -> str:
        if not totals:
            return "<i>(Tidak ada agregasi yang ditulis.)</i>"

        months = sorted(totals.keys())
        lines, grand = [], {}

        def _fmt_num(x):
            try:
                return f"{x:,.0f}"
            except Exception:
                return str(x)

        from app.mapping import MASTER_KEYS  # dipakai hanya saat ringkas
        for m in months:
            d = totals[m] or {}
            parts = []
            for k in MASTER_KEYS:
                if k in d and d[k]:
                    parts.append(f"{k}={_fmt_num(d[k])}")
                    grand[k] = grand.get(k, 0) + (d[k] or 0)
            lines.append(f"<code>{m}</code> | " + ", ".join(parts))

        gparts = [f"{k}={_fmt_num(v)}" for k, v in grand.items() if v]
        if gparts:
            lines.append("<b>Total</b> • " + ", ".join(gparts))
        return "<br/>".join(lines)


# =========================================================
# Wiring utama
# =========================================================
def wire_browse_handlers(parent: QtWidgets.QWidget, ui):
    # ATAS = BARGE
    if hasattr(ui, "pushButton_2") and hasattr(ui, "fileLineEdit"):
        def _browse_barge():
            path = _browse_into(parent, ui.fileLineEdit, "Pilih Barge workbook")
            _populate_sheet_combo(ui, path)  # combobox diisi dari file barge
        ui.pushButton_2.clicked.connect(_browse_barge)

    # BAWAH = MASTER
    if hasattr(ui, "pushButton_3") and hasattr(ui, "fileLineEdit_2"):
        def _browse_master():
            _browse_into(parent, ui.fileLineEdit_2, "Pilih FLF Report Data (master)")
        ui.pushButton_3.clicked.connect(_browse_master)

    # tombol Submit → proses
    if hasattr(ui, "pushButton"):
        def _on_submit():
            f_barge  = ui.fileLineEdit.text().strip()    if hasattr(ui, "fileLineEdit")   else ""
            f_master = ui.fileLineEdit_2.text().strip()  if hasattr(ui, "fileLineEdit_2") else ""

            if not f_master or not f_master.lower().endswith(".xlsx") or not os.path.exists(f_master):
                QtWidgets.QMessageBox.warning(parent, "Master bermasalah",
                    "Pilih file Master .xlsx yang valid.")
                return
            if not f_barge or not f_barge.lower().endswith(".xlsx") or not os.path.exists(f_barge):
                QtWidgets.QMessageBox.warning(parent, "Barge bermasalah",
                    "Pilih file Barge .xlsx yang valid.")
                return

            # ==== Ambil parameter dari UI utama ====
            year = None
            if hasattr(ui, "comboYear"):
                try:
                    year = int(ui.comboYear.currentText())
                except Exception:
                    pass
            if year is None:
                year = max(MASTER_SHEET_NAMES.keys())  # fallback

            sheet = ui.comboSheet.currentText() if hasattr(ui, "comboSheet") else DEFAULT_BARGE_SHEET
            start_row = ui.spinStart.value() if hasattr(ui, "spinStart") else 246
            row_count = ui.spinCount.value() if hasattr(ui, "spinCount") else 29
            only_completed = ui.chkCompleted.isChecked() if hasattr(ui, "chkCompleted") else True

            # >>> SELALU APPLY (tanpa dry-run)
            dry_run = False

            # Preflight
            try:
                _preflight_barge(f_barge, sheet, start_row)
            except Exception as e:
                _show_error_dialog(parent, str(e))
                return

            opts = RunOptions(
                master_path=f_master,
                barge_path=f_barge,
                barge_sheet=sheet,
                target_year=year,
                start_row=start_row,
                row_count=row_count,
                only_completed=only_completed,
                dry_run=dry_run,
            )

            # Jalankan via StartDialog (tampilkan ringkasan plain text)
            def _runner():
                return run_pipeline(opts)  # -> (totals, logs)

            preface = _options_summary_text(opts)
            StartDialog(parent, _runner, is_dry_run=opts.dry_run, preface_text=preface).exec_()

        ui.pushButton.clicked.connect(_on_submit)
