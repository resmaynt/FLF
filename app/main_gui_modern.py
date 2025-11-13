# app/main_gui_modern.py — PyQt5 (Refactor: ramah pemula)
from __future__ import annotations
from dataclasses import dataclass
from typing import Callable, Iterable, List, Tuple

from PyQt5 import QtWidgets, QtCore, QtGui
from functools import partial
# Pastikan resources.qrc sudah di-compile menjadi ui/resources_rc.py
# Contoh: pyrcc5 resources.qrc -o ui/resources_rc.py
import ui.resources_rc  # noqa: F401
from app.popup import confirm_summary

# Import dari modul Anda sendiri
from app.config import MASTER_SHEET_NAMES, DEFAULT_BARGE_SHEET, RunOptions
from app.main_logic import run_pipeline

from openpyxl import load_workbook
import os
import inspect

# === Tambahan untuk reset runtime saat End ===
import importlib
import app.config as cfg
import app.mapping as mapping
import app.main_logic as logic

def reset_runtime_state() -> None:
    """Reload modul & bersihkan override volatile supaya state tidak nempel antar-run."""
    try:
        importlib.reload(cfg)
        importlib.reload(mapping)
        importlib.reload(logic)
        if hasattr(cfg, "FORCE_MONTH_OVERRIDE"):
            cfg.FORCE_MONTH_OVERRIDE = ""
    except Exception:
        pass

# =============================
# Konstanta UI untuk konsistensi
# =============================
SIDEBAR_WIDTH = 220
WINDOW_MIN_W = 1100
WINDOW_MIN_H = 640

LINEEDIT_HEIGHT = 45
BUTTON_WIDTH = 90
CTRL_HEIGHT = 45   # ← match the line edit / button height

# Opsi panel
YEAR_WIDTH = 125
SHEET_WIDTH = 110
START_WIDTH = 80
COUNT_WIDTH = 80



# Padding/Spacing
TOP_PAD_FORM = 1
GAP_AFTER_FILES = 1
OPT_TOP_MARGIN = 20
OPT_RIGHT_MARGIN_EXTRA = BUTTON_WIDTH  # supaya rata dengan tombol Browse
OPT_VSPACE = 10
OPT_HSPACE = 16 # jarak antar kotak dalam satu baris


# =============================
# Utilitas kecil
# =============================

def set_nav_selected(btn: QtWidgets.QPushButton, selected: bool) -> None:
    """Tandai tombol sidebar sedang terpilih (untuk QSS)."""
    btn.setChecked(selected)
    btn.setProperty("selected", selected)
    btn.style().unpolish(btn)
    btn.style().polish(btn)


def browse_excel(parent: QtWidgets.QWidget, title: str) -> str:
    """Dialog pilih file Excel; mengembalikan path atau string kosong."""
    start = (
        QtCore.QStandardPaths.writableLocation(QtCore.QStandardPaths.DocumentsLocation)
        or os.path.expanduser("~")
    )
    path, _ = QtWidgets.QFileDialog.getOpenFileName(
        parent, title, start, "Excel (*.xlsx);;All Files (*.*)"
    )
    return path or ""


def read_sheetnames(xlsx_path: str) -> List[str]:
    """Ambil daftar sheet dari workbook; jika gagal, kembalikan list kosong."""
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        return list(wb.sheetnames)
    except Exception:
        return []


# =============================
# Worker thread
# =============================
class Worker(QtCore.QThread):
    """Thread untuk menjalankan pipeline agar UI tidak freeze."""

    progress = QtCore.pyqtSignal(str)
    finished = QtCore.pyqtSignal(object, object)  # totals, logs
    failed = QtCore.pyqtSignal(str)

    def __init__(self, runner: Callable[..., Tuple[object, object]], parent=None) -> None:
        super().__init__(parent)
        self._runner = runner

    def run(self) -> None:  # type: ignore[override]
        try:
            # Jika runner mendukung argumen progress, berikan sinyal progress.emit
            if "progress" in inspect.signature(self._runner).parameters:
                totals, logs = self._runner(progress=self.progress.emit)
            else:
                totals, logs = self._runner()
            self.finished.emit(totals, logs)
        except Exception as e:
            self.failed.emit(str(e))


def done_popup(parent: QtWidgets.QWidget, ok: bool = True) -> None:
    mb = QtWidgets.QMessageBox(parent)
    mb.setWindowTitle("Selesai" if ok else "Error")
    mb.setIcon(QtWidgets.QMessageBox.Information if ok else QtWidgets.QMessageBox.Critical)
    mb.setText("Proses selesai!" if ok else "Proses gagal.")
    mb.exec_()


# =============================
# Komponen background penuh jendela
# =============================
class BackgroundLabel(QtWidgets.QLabel):
    """Label yang menampilkan gambar latar dan otomatis di-scale."""

    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self._pix: QtGui.QPixmap | None = None
        self.setScaledContents(False)

    def set_image(self, path: str) -> None:
        pix = QtGui.QPixmap(path)
        if not pix.isNull():
            self._pix = pix
            self._update()

    def resizeEvent(self, e) -> None:  # type: ignore[override]
        super().resizeEvent(e)
        self._update()

    def _update(self) -> None:
        if self._pix:
            self.setPixmap(
                self._pix.scaled(
                    self.size(),
                    QtCore.Qt.KeepAspectRatioByExpanding,
                    QtCore.Qt.SmoothTransformation,
                )
            )


# =============================
# Halaman Form → Run/Log (factory)
# =============================

def build_flf_page(kind_label: str, runner_func: Callable[..., Tuple[object, object]], key: str) -> QtWidgets.QWidget:
    page = QtWidgets.QWidget(objectName=f"{key}Page")
    page.setAttribute(QtCore.Qt.WA_StyledBackground, True)

    #layout root
    root_v = QtWidgets.QVBoxLayout(page)
    root_v.setContentsMargins(48, 130, 48, 16)   # ⇠ top margin kecil
    root_v.setSpacing(10)

    # Judul halaman (tetap)
    title = QtWidgets.QLabel(kind_label)
    title.setProperty("class", "h1")
    title.setAlignment(QtCore.Qt.AlignHCenter)

    header = QtWidgets.QWidget()
    header.setFixedHeight(180)  # tinggi tetap untuk header
    hb = QtWidgets.QVBoxLayout(header)
    hb.setContentsMargins(0, 100, 0, 24)  # jarak dalam header saja
    hb.addWidget(title, 0, QtCore.Qt.AlignHCenter)

    # Stack: 0=Form, 1=Run/Log
    root_v.setContentsMargins(24, 24, 24, 16)  # margin normal, bukan 130
    root_v.addWidget(header, 0)  # header tidak “mendorong” form
    stack = QtWidgets.QStackedWidget()
    root_v.addWidget(stack, 1)

    # ---------- FORM ----------
    form = QtWidgets.QWidget(objectName=f"{key}Form")
    form.setAttribute(QtCore.Qt.WA_StyledBackground, True)

    vbox = QtWidgets.QVBoxLayout(form)
    vbox.setContentsMargins(0, 10, 0, 0)
    vbox.setSpacing(16)                   # was: 6
    vbox.setAlignment(QtCore.Qt.AlignTop) # paksa konten nempel ke atas

    # Baris input file
    row_barge = QtWidgets.QHBoxLayout()
    row_master = QtWidgets.QHBoxLayout()

    lbl_barge = QtWidgets.QLabel("FLF File:", objectName="lblBarge")
    lbl_master = QtWidgets.QLabel("Final File:", objectName="lblMaster")
    for lbl in (lbl_barge, lbl_master):
        lbl.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        lbl.setProperty("class", "formLabel")

    le_barge = QtWidgets.QLineEdit(placeholderText="FLF Report Data", objectName="fileLineEdit")
    le_master = QtWidgets.QLineEdit(placeholderText="Draft Power BI ", objectName="fileLineEdit_2")

    btn_barge = QtWidgets.QPushButton("Browse", objectName="pushButton_2"); btn_barge.setProperty("class", "primary")
    btn_master = QtWidgets.QPushButton("Browse", objectName="pushButton_3"); btn_master.setProperty("class", "primary")

    for w in (le_barge, le_master):
        w.setMinimumHeight(LINEEDIT_HEIGHT)
        w.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
    for b in (btn_barge, btn_master):
        b.setMinimumHeight(LINEEDIT_HEIGHT)

    label_w = max(lbl_barge.sizeHint().width(), lbl_master.sizeHint().width())
    lbl_barge.setFixedWidth(label_w)
    lbl_master.setFixedWidth(label_w)

    row_barge.addWidget(lbl_barge, 0, QtCore.Qt.AlignVCenter)
    row_barge.addWidget(le_barge, 1, QtCore.Qt.AlignVCenter)
    row_barge.addWidget(btn_barge, 0, QtCore.Qt.AlignVCenter)
    row_barge.setStretch(1, 1)

    row_master.addWidget(lbl_master, 0, QtCore.Qt.AlignVCenter)
    row_master.addWidget(le_master, 1, QtCore.Qt.AlignVCenter)
    row_master.addWidget(btn_master, 0, QtCore.Qt.AlignVCenter)
    row_master.setStretch(1, 1)

    vbox.addLayout(row_barge)
    vbox.addLayout(row_master)

    # ---------- Options ----------
    options = QtWidgets.QGroupBox(objectName="optionsPanel")
    options.setTitle("")
    grid = QtWidgets.QGridLayout(options)
    grid.setContentsMargins(0, 0, BUTTON_WIDTH + row_barge.spacing(), 0)
    grid.setHorizontalSpacing(OPT_HSPACE)
    grid.setVerticalSpacing(OPT_VSPACE)

    row_sheet = QtWidgets.QHBoxLayout()
    lbl_sheet = QtWidgets.QLabel("Sheet:")
    lbl_sheet.setProperty("class", "formLabel")
    lbl_sheet.setFixedWidth(label_w)
    lbl_sheet.setContentsMargins(0, 0, 0, 0)
    row_sheet.addWidget(lbl_sheet, 0, QtCore.Qt.AlignVCenter)
    row_sheet.addWidget(options, 1, QtCore.Qt.AlignVCenter)
    row_sheet.setSpacing(12)

    # combo_year = QtWidgets.QComboBox()
    # for y in sorted(MASTER_SHEET_NAMES.keys()):
    #     combo_year.addItem(str(y))
    # combo_year.setCurrentText(str(max(MASTER_SHEET_NAMES.keys())))
    
    year_edit = QtWidgets.QLineEdit()
    year_edit.setPlaceholderText("mis. 2026")
    year_edit.setAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)


    year_edit.setText(str(max(MASTER_SHEET_NAMES.keys())))

    combo_sheet = QtWidgets.QComboBox(); combo_sheet.addItem(DEFAULT_BARGE_SHEET)
    spin_start = QtWidgets.QSpinBox();  spin_start.setRange(1, 1_000_000); spin_start.setValue(246)
    spin_count = QtWidgets.QSpinBox();  spin_count.setRange(0, 1_000_000); spin_count.setValue(29)

    # Hilangkan tombol panah ↑↓
    spin_start.setButtonSymbols(QtWidgets.QAbstractSpinBox.NoButtons)
    spin_count.setButtonSymbols(QtWidgets.QAbstractSpinBox.NoButtons)

    # (Opsional) Matikan scroll wheel biar nggak keubah tanpa sengaja
    def _no_wheel(e): e.ignore()
    spin_start.wheelEvent = _no_wheel
    spin_count.wheelEvent = _no_wheel


    def cell_top(caption: str, ctrl: QtWidgets.QWidget, width: int) -> QtWidgets.QWidget:
        w = QtWidgets.QWidget()
        v = QtWidgets.QVBoxLayout(w)
        v.setContentsMargins(0, 0, 0, 0)
        v.setSpacing(6)
        lab = QtWidgets.QLabel(caption); lab.setAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
        lab.setProperty("class", "optCaption")
        ctrl.setFixedSize(width, CTRL_HEIGHT)
        ctrl.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        v.addWidget(lab); v.addWidget(ctrl)
        return w

    grid.addWidget(cell_top("Year",        year_edit,  YEAR_WIDTH),  0, 0)
    grid.addWidget(cell_top("Barge Sheet", combo_sheet, SHEET_WIDTH), 0, 1)
    grid.addWidget(cell_top("Row Count",   spin_count,  COUNT_WIDTH), 0, 2)
    grid.addWidget(cell_top("Start Row",   spin_start,  START_WIDTH), 0, 3)
    grid.setColumnStretch(0, 1); grid.setColumnStretch(1, 1)

    # === CHECKBOX BARU ===
    cb_clear = QtWidgets.QCheckBox("Clear row before write")
    cb_clear.setChecked(True)
    cb_clear.setFixedHeight(CTRL_HEIGHT)

    grid.addWidget(cb_clear, 1, 0, 1, 2)

    vbox.addLayout(row_sheet)

    # ---------- Submit ----------
    btn_submit = QtWidgets.QPushButton("Submit", objectName="pushButton")
    btn_submit.setProperty("class", "primary")
    btn_submit.setMinimumHeight(LINEEDIT_HEIGHT)
    btn_submit.setMinimumWidth(BUTTON_WIDTH)

    submit_row = QtWidgets.QHBoxLayout()
    submit_row.setContentsMargins(0, 16, 0, 0)
    submit_row.setAlignment(QtCore.Qt.AlignHCenter)
    submit_row.addWidget(btn_submit)
    vbox.addLayout(submit_row)
    stack.addWidget(form)

    # ---------- RUN/LOG ----------
    run = QtWidgets.QWidget(objectName=f"{key}Run")
    run.setAttribute(QtCore.Qt.WA_StyledBackground, True)
    run_v = QtWidgets.QVBoxLayout(run); run_v.setSpacing(12)
    run_v.setContentsMargins(0, 0, 0, 0)
    run_v.setAlignment(QtCore.Qt.AlignTop)
    power_btn = QtWidgets.QPushButton()

    power_btn.setObjectName("powerBtn")
    power_btn.setIcon(QtGui.QIcon(":/img/power.ico"))   # pastikan ada di resources
    power_btn.setIconSize(QtCore.QSize(120, 120))       # ukuran ikon
    power_btn.setFixedSize(200, 200)                    # ukuran tombol
    power_btn.setFlat(True)
    power_btn.setToolTip("")                            # hilangkan tooltip
    power_btn.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
    power_btn.setFocusPolicy(QtCore.Qt.NoFocus)         # tanpa outline fokus
    run_v.addWidget(power_btn, 0, QtCore.Qt.AlignHCenter | QtCore.Qt.AlignTop)

    # Hapus tombol Start; End jadi biru
    log = QtWidgets.QTextEdit(objectName="logArea"); log.setReadOnly(True)
    run_v.addWidget(log, 1)

    end_row = QtWidgets.QHBoxLayout(); end_row.addStretch(1)
    btn_end = QtWidgets.QPushButton("End"); 
    btn_end.setObjectName("pushButton")
    btn_end.setEnabled(False); 
    btn_end.setProperty("class", "primary")
    end_row.addWidget(btn_end)
    run_v.addLayout(end_row)

    stack.addWidget(run)

    # --- Simple page switch that toggles header visibility
    def switch_to_form():
        header.setVisible(True)       # header tampil di Form
        stack.setCurrentIndex(0)

    def switch_to_run():
        header.setVisible(False)      # header disembunyikan di Run
        stack.setCurrentIndex(1)

    # ===== Handlers =====
    def set_path(line_edit: QtWidgets.QLineEdit, title: str) -> None:
        path = browse_excel(page, title)
        if path:
            line_edit.setText(path)

    def fill_sheets(combo: QtWidgets.QComboBox, path: str) -> None:
        combo.clear()
        sheets = read_sheetnames(path)
        combo.addItems(sheets or [DEFAULT_BARGE_SHEET])
        if sheets and DEFAULT_BARGE_SHEET in sheets:
            combo.setCurrentText(DEFAULT_BARGE_SHEET)


    def on_pick_barge():
        set_path(le_barge, "Pilih Barge workbook")
        fill_sheets(combo_sheet, le_barge.text())
    btn_barge.clicked.connect(on_pick_barge)

    btn_master.clicked.connect(lambda: set_path(le_master, "Pilih FLF Report Data (master)"))


    def build_options() -> RunOptions:
        # parse tahun dari year_edit (fallback ke nilai default jika kosong)
        txt = year_edit.text().strip()
        try:
            yr = int(txt)
        except Exception:
            yr = max(MASTER_SHEET_NAMES.keys())  # fallback aman

        return RunOptions(
            master_path=le_master.text().strip(),
            barge_path=le_barge.text().strip(),
            barge_sheet=combo_sheet.currentText().strip() or DEFAULT_BARGE_SHEET,
            target_year=yr,
            start_row=spin_start.value(),
            row_count=spin_count.value(),
            only_completed=True,
            dry_run=False,
            clear_before_write=cb_clear.isChecked(),
        )


    def validate_paths(master: str, barge: str) -> bool:
        if not (master and master.lower().endswith(".xlsx") and os.path.exists(master)):
            QtWidgets.QMessageBox.warning(page, "Master bermasalah", "Pilih file Master .xlsx yang valid.")
            return False
        if not (barge and barge.lower().endswith(".xlsx") and os.path.exists(barge)):
            QtWidgets.QMessageBox.warning(page, "Barge bermasalah", "Pilih file Barge .xlsx yang valid.")
            return False
        return True

    def on_submit() -> None:
        master, barge = le_master.text().strip(), le_barge.text().strip()
        if not validate_paths(master, barge):
            return
        opts = build_options()
        summary = (
            f"Master       : {opts.master_path}\n"
            f"Barge        : {opts.barge_path}\n"
            f"Year         : {opts.target_year}\n"
            f"Barge Sheet  : {opts.barge_sheet}\n"
            f"Start Row    : {opts.start_row}\n"
            f"Row Count    : {opts.row_count}\n"
            f"Clear Row?   : {opts.clear_before_write}\n"
        )

        if confirm_summary(page, summary):
            switch_to_run()
            log.clear()
            btn_end.setEnabled(False)
            page._opts = opts  # simpan opsi di page
    btn_submit.clicked.connect(on_submit)

    def runner_builder() -> Callable[..., Tuple[object, object]]:
        opts: RunOptions | None = getattr(page, "_opts", None)
        if not opts:
            raise RuntimeError("Options belum di-submit.")
        # Pakai partial, jangan lambda
        return partial(runner_func, opts)

    def on_start() -> None:
        power_btn.setEnabled(False)
        log.append("Processing data...")
        worker = Worker(runner_builder(), page)
        page._worker = worker
        worker.progress.connect(lambda s: log.append(s))

        def done(totals, logs):
            if logs:
                log.append("")          # baris kosong pemisah
                log.append("Details")   # tidak bold, tidak berwarna
                for line in logs:
                    s = str(line)
                    up = s.upper()
                    if "ERROR" in up or "FAILED" in up:
                        # hanya error yang merah
                        log.append(f"<span style='color:#b91c1c;'>{s}</span>")
                    else:
                        # baris normal: tanpa warna/format
                        log.append(s)

            log.append("Execution completed.")  # tidak bold/warna
            done_popup(page, ok=True)
            btn_end.setEnabled(True)

        worker.finished.connect(done)
        worker.failed.connect(lambda e: (done_popup(page, ok=False), btn_end.setEnabled(True)))
        worker.start()

    # klik ikon power = mulai
    power_btn.clicked.connect(on_start)

    # End → reset runtime, balik ke form, bersihkan log, hidupkan ikon power lagi
    def on_end_clicked() -> None:
        # putus worker
        if hasattr(page, "_worker"):
            page._worker = None
        # reset modul/state & opsi
        reset_runtime_state()
        if hasattr(page, "_opts"):
            page._opts = None
        # clear form agar submit ulang → state fresh
        le_barge.clear()
        le_master.clear()
        combo_sheet.clear()
        combo_sheet.addItem(DEFAULT_BARGE_SHEET)
        switch_to_form()
        power_btn.setEnabled(True)
        log.clear()

    btn_end.clicked.connect(on_end_clicked)

    return page

# =============================
# Bangun jendela utama (QDialog)
# =============================

def create_window() -> QtWidgets.QDialog:
    dlg = QtWidgets.QDialog()
    dlg.setWindowTitle("FLF Automation System")
    dlg.resize(WINDOW_MIN_W, WINDOW_MIN_H)
    dlg.setMinimumSize(WINDOW_MIN_W, WINDOW_MIN_H)

    # pakai logo-side.jpg di title bar kiri atas
    dlg.setWindowIcon(QtGui.QIcon(":/img/logo-side.jpg"))

    # === Tambahan: hilangkan tanda '?' dan tampilkan tombol minimize
    dlg.setWindowFlag(QtCore.Qt.WindowContextHelpButtonHint, False)
    dlg.setWindowFlag(QtCore.Qt.WindowMinimizeButtonHint, True)
    dlg.setWindowFlag(QtCore.Qt.WindowSystemMenuHint, True)
    #dlg.setWindowFlag(QtCore.Qt.WindowMaximizeButtonHint, False)

    # Background bersama
    root_bg = BackgroundLabel(dlg)
    root_bg.setObjectName("RootBg")
    root_bg.setGeometry(dlg.rect())
    root_bg.lower()

    # Sidebar transparan + tint
    side = QtWidgets.QFrame(dlg)
    side.setObjectName("SideBar")
    side.setGeometry(QtCore.QRect(0, 0, SIDEBAR_WIDTH, dlg.height()))
    side.setAttribute(QtCore.Qt.WA_StyledBackground, True)

    side_tint = QtWidgets.QFrame(side)
    side_tint.setGeometry(0, 0, SIDEBAR_WIDTH, dlg.height())
    side_tint.setStyleSheet("background: rgba(0,0,0,0.30);")
    side_tint.lower()

    brand = QtWidgets.QLabel(side)
    brand.setPixmap(QtGui.QPixmap(":/img/logo.png"))
    brand.setScaledContents(True)
    brand.setGeometry(66, 12, 70, 46)

    cap = QtWidgets.QLabel("PT Indo Tambangraya Megah Tbk", side)
    cap.setObjectName("brandCaption")
    cap.setAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
    cap.setGeometry(20, 64, 180, 20)

    nav_home  = QtWidgets.QPushButton("  Home", side);        nav_home.setObjectName("navHome")
    nav_month = QtWidgets.QPushButton("  FLF Monthly", side); nav_month.setObjectName("navMonthly")

    for i, b in enumerate((nav_home, nav_month)):
        b.setCheckable(True)
        b.setGeometry(12, 150 + i * 56, 196, 40)

    for w in (brand, cap, nav_home, nav_month):
        w.raise_()

    # Area kanan (stack halaman)
    pages = QtWidgets.QStackedWidget(dlg)
    pages.setGeometry(QtCore.QRect(SIDEBAR_WIDTH, 0, dlg.width() - SIDEBAR_WIDTH, dlg.height()))
    pages.setObjectName("MainArea")
    pages.setAttribute(QtCore.Qt.WA_StyledBackground, True)

    # Halaman Home (kosong; transparan, pakai root_bg)
    home = QtWidgets.QWidget(objectName="HomePage")
    home.setAttribute(QtCore.Qt.WA_StyledBackground, False)
    _ = QtWidgets.QVBoxLayout(home)

    # Hanya Monthly
    monthly = build_flf_page("FLF Monthly", run_pipeline, key="Monthly")

    for p in (home, monthly):
        pages.addWidget(p)

    def set_root_bg(where: str) -> None:
        root_bg.set_image(":/img/bg.png" if where == "home" else ":/img/bg2.png")

    set_root_bg("home")

    # Navigation
    nav_home.clicked.connect(lambda: (set_nav_selected(nav_home, True),
                                      set_nav_selected(nav_month, False),
                                      pages.setCurrentIndex(0),
                                      set_root_bg("home")))
    nav_month.clicked.connect(lambda: (set_nav_selected(nav_home, False),
                                       set_nav_selected(nav_month, True),
                                       pages.setCurrentIndex(1),
                                       set_root_bg("monthly")))

    set_nav_selected(nav_home, True)
    pages.setCurrentIndex(0)

    # Responsif pada resize
    _old_resize = dlg.resizeEvent
    def on_resize(e) -> None:
        if _old_resize:
            QtWidgets.QDialog.resizeEvent(dlg, e)
        root_bg.setGeometry(dlg.rect())
        side.setGeometry(0, 0, SIDEBAR_WIDTH, dlg.height())
        side_tint.setGeometry(0, 0, SIDEBAR_WIDTH, dlg.height())
        pages.setGeometry(SIDEBAR_WIDTH, 0, dlg.width() - SIDEBAR_WIDTH, dlg.height())
    dlg.resizeEvent = on_resize

    return dlg


# =============================
# Entry point opsional (untuk test cepat)
# =============================
if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    # Terapkan stylesheet Anda di sini jika perlu: lihat theme.qss
    w = create_window()
    w.show()
    sys.exit(app.exec_())
